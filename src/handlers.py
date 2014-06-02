import zook
import tornado.websocket
import uuid
import time
import ujson as json
import codecs
import os
import sys
import traceback
import re
import openpyxl
from collections import OrderedDict
import decimal


class ClientHandler(tornado.web.RequestHandler):
    """docstring for ClientHandler"""
    def get(self):
        path = os.path.join(self.application.public_path, 'client.html')
        with codecs.open(path, 'r', 'utf-8') as f:
            read_data = f.read()
        self.write(read_data)


class ServerHandler(tornado.web.RequestHandler):
    """docstring for ServerHandler"""
    def get(self):
        path = os.path.join(self.application.public_path, 'server.html')
        with codecs.open(path, 'r', 'utf-8') as f:
            read_data = f.read()
        self.write(read_data)


class ExportHandler(tornado.web.RequestHandler):
    """docstring for ExportHandler"""
    def get(self):
        key = self.get_argument('key')
        self.render(key)
        file_name = 'session-' + key + '.xlsx'
        path = os.path.join(self.application.data_path, file_name)
        buf_size = 4096
        self.set_header('Content-Type', 'application/octet-stream')
        self.set_header('Content-Disposition', 'attachment; filename=' + file_name)
        with open(path, 'rb') as f:
            while True:
                data = f.read(buf_size)
                if not data:
                    break
                self.write(data)
        self.finish()

    def render(self, key):
        wb = openpyxl.Workbook(encoding='utf-8')
        ws = wb.active
        ws.title = 'Session'
        path_json = os.path.join(self.application.data_path, 'session-' + key + '.json')
        path_xlsx = os.path.join(self.application.data_path, 'session-' + key + '.xlsx')
        data = None
        with open(path_json, 'r') as f:
            data = json.load(f)
        row = 0
        headers = (
            'Phase',
            'Period',
            'Group',
            'Subject',
            'UnitCost',
            'QuantityInitial',
            'QuantityReached',
            'Direction',
            'UpCovered',
            'DownCovered',
            'CoinFlip',
            'Outcome',
            'Role',
            'Provide',
            'Bid',
            'Ask',
            'PeriodStartingBalance',
            'PhaseStartingBalance',
            'Balance',
            'PeriodProfit',
            'PhaseProfit',
            'TotalProfit'
        )
        col = 0
        for h in headers:
            cell = ws.cell(row=row, column=col)
            cell.value = h
            cell.style.font.bold = True
            # ws.column_dimensions[openpyxl.cell.get_column_letter(col + 1)].width = 15
            col += 1
        row += 1
        col = 0
        phases = OrderedDict(sorted(data['phases'].items(), key=lambda t: int(t[0])))
        for i, ph in enumerate(phases.values()):
            periods = OrderedDict(sorted(ph['periods'].items(), key=lambda t: int(t[0])))
            for j, pe in enumerate(periods.values()):
                groups = OrderedDict(sorted(pe['groups'].items(), key=lambda t: int(t[0])))
                for k, g in enumerate(groups.values()):
                    subjects = OrderedDict(sorted(g['subjects'].items(), key=lambda t: t[1]['name']))
                    for s in subjects.values():
                        col = 0
                        cells = self.generate_session_row(ph, pe, g, s)
                        for c in cells:
                            ws.cell(row=row, column=col).value = c
                            col += 1
                        row += 1
        ws = wb.create_sheet()
        ws.title = 'Subjects'
        headers = (
            'Subject',
            'Suspended',
            'Robot',
            'Balance',
            'TotalProfit',
            'AmountToPay',
            'RealName',
            'IdentificationNumber',
            'Address',
            'PostalCode',
            'Location',
            'Email'
        )
        row = 0
        col = 0
        for h in headers:
            cell = ws.cell(row=row, column=col)
            cell.value = h
            cell.style.font.bold = True
            col += 1
        row = 1
        for s in data['subjects']:
            col = 0
            r = (
                s['name'],
                s['is_suspended'],
                s['is_robot'],
                s['current_balance'],
                s['total_profit'],
                s['amount_to_pay'],
                s['real_name'],
                s['identification_number'],
                s['address'],
                s['postal_code'],
                s['location'],
                s['email']
            )
            for c in r:
                ws.cell(row=row, column=col).value = c
                col += 1
            row += 1
        wb.save(path_xlsx)

    def set_cell_value(self, ws, row, column, value, centered=False):
        c = ws.cell(row=row, column=column)
        if centered:
            c.style.alignment.horizontal = 'center'
            c.style.alignment.vertical = 'center'
        c.value = value

    def generate_session_row(self, ph, pe, g, s):
        role = None
        provide = None
        bid = None
        ask = None
        period_profit = None
        period_balance = None
        phase_profit = None
        phase_balance = None
        if s['key'] in g['roles']:
            role = g['roles'][s['key']]
        if s['key'] in g['provides']:
            provide = g['provides'][s['key']]
        if s['key'] in g['bids']:
            bid = g['bids'][s['key']]
        if s['key'] in g['asks']:
            ask = g['asks'][s['key']]
        if s['key'] in pe['balances']:
            period_balance = pe['balances'][s['key']]
        if s['key'] in pe['profits']:
            period_profit = pe['profits'][s['key']]
        if s['key'] in ph['balances']:
            phase_balance = ph['balances'][s['key']]
        if s['key'] in ph['profits']:
            phase_profit = ph['profits'][s['key']]
        row = (
            ph['key'],
            pe['key'],
            g['key'],
            s['name'],
            pe['cost'],
            g['quantity_initial'],
            g['quantity_reached'],
            g['direction'],
            g['up_covered'],
            g['down_covered'],
            g['coin_flip'],
            g['outcome'],
            role,
            provide,
            bid,
            ask,
            period_balance,
            phase_balance,
            s['current_balance'],
            period_profit,
            phase_profit,
            s['total_profit']
        )
        return row


class InvalidOperationException(Exception):
    pass


class UnauthorizedOperationException(Exception):
    pass


class InvalidMessageException(Exception):
    pass


class SocketHandler(tornado.websocket.WebSocketHandler):
    """docstring for SocketHandler"""
    def open(self):
        self.is_open = True
        self.key = str(uuid.uuid4())
        self.application.sockets[self.key] = self
        self.session = self.find_session()
        self.is_initialized = False
        self.is_experimenter = False
        self.timer = None
        self.timer_started_at = None

    def on_close(self):
        self.is_open = False
        if hasattr(self, 'key'):
            del self.application.sockets[self.key]
            s = self.application.get_subject(self.key)
            if s is not None:
                s.set_state('dropped')
                self.notify('set_subject', s)

    def on_message(self, message):
        o = None
        id = None
        message_type = None
        reply = None
        try:
            try:
                o = json.loads(message)
            except:
                raise InvalidMessageException(
                    'Message should be in json format')
            if 'id' not in o:
                raise InvalidMessageException(
                    'Message should contain "id"')
            id = o['id']
            if 'type' not in o:
                raise InvalidMessageException(
                    'Message should contain "type"')
            message_type = o['type']
            if self.is_initialized is False and message_type != 'initialize':
                raise InvalidOperationException(
                    'socket should be initialized first')
            elif message_type == 'initialize':
                reply = self.init(o)
            elif message_type == 'get_session':
                reply = self.get_session(o)
            elif message_type == 'set_session':
                reply = self.set_session(o)
            elif message_type == 'get_group':
                reply = self.get_group(o)
            elif message_type == 'get_subject':
                reply = self.get_subject(o)
            elif message_type == 'set_subject':
                reply = self.set_subject(o)
            elif message_type == 'get_subjects':
                reply = self.get_subjects(o)
            elif message_type == 'authorize':
                reply = self.authorize(o)
            elif message_type == 'suspend_subject':
                reply = self.suspend_subject(o)
            elif message_type == 'start_session':
                reply = self.start_session(o)
            elif message_type == 'pause_session':
                reply = self.pause_session(o)
            elif message_type == 'resume_session':
                reply = self.resume_session(o)
            elif message_type == 'stop_session':
                reply = self.stop_session(o)
            elif message_type == 'continue_session':
                reply = self.continue_session(o)
            elif message_type == 'skip_phase':
                reply = self.skip_phase(o)
            elif message_type == 'reset':
                reply = self.reset(o)
            else:
                raise InvalidOperationException('Unknown message type')
            self.send(message_type, reply, id)
        except InvalidMessageException as e:
            self.send('invalid_message', str(e), id)
        except InvalidOperationException as e:
            self.send('invalid_operation', str(e), id)
        except UnauthorizedOperationException as e:
            self.send('unauthorized_operation', str(e), id)
        except Exception as e:
            self.send('internal_error', traceback.format_exc(), id)
        else:
            self.notify(message_type, reply)

    def send(self, message_type='reply', message=None, id=None):
        m = dict(
            type=message_type,
            timestamp=int(time.time() * 1000),
            data=message
        )
        if id is not None:
            m['id'] = id
        js = json.dumps(m)
        self.write_message(js)
        del m
        del js

    def notify(self, message_type, message, is_global=False):
        ignore_list = (
            'initialize',
            'authorize',
            'get_session'
            )
        if message_type not in ignore_list:
            for e in self.session.experimenters.values():
                socket = self.application.get_socket(e.key)
                if socket is not None and socket.is_open and e.key is not self.key:
                    ses = self.application.clone_session(self.session)
                    socket.send('get_session', ses)

    @staticmethod
    def check_data(message, keys=[]):
        if 'data' not in message or message['data'] is None:
            raise InvalidOperationException('Message should contain "data"')
        if len(keys) > 0:
            data = message['data']
            for key in keys:
                if key not in data or not data[key] or not data[key].strip():
                    raise InvalidOperationException(
                        'Message "data" should contain a "{0}"'.format(key)
                        )

    def check_experimenter(self):
        if not self.is_experimenter:
            raise UnauthorizedOperationException('Access denied')

    def find_session(self):
        session = None
        if len(self.application.sessions) > 0:
            session = next(
                (s for s in self.application.sessions.values()),
                None
                )
        if session is None:
            session = zook.Session()
            self.application.sessions[session.key] = session
        return session

    def parse_key(self, data):
        try:
            key = uuid.UUID(data)
            return str(key)
        except:
            return None

    def is_numeric(self, value):
        return value is not None and re.match('^\d+(\.\d+){0,1}$', str(value)) is not None

    def init(self, message):
        data = {}
        subject = None
        if 'data' in message and message['data'] is not None:
            key = self.parse_key(message['data'])
            if key is not None:
                socket = self.application.get_socket(key)
                if socket is not None and socket.is_open:
                    socket.close()
                del self.application.sockets[self.key]
                self.key = key
                self.application.sockets[self.key] = self
                exp = self.application.get_experimenter(self.key)
                if exp is not None:
                    self.is_experimenter = True
                    self.session = exp.session
                else:
                    sub = self.application.get_subject(self.key)
                    if sub is not None:
                        self.session = sub.session
                        if zook.Subject.states[sub.state] == 'dropped':
                            sub.restore_state()
                    else:
                        self.session = self.find_session()

        data['key'] = self.key
        data['session'] = self.application.clone_session(self.session)
        data['is_experimenter'] = self.is_experimenter
        self.is_initialized = True
        return data

    def get_session(self, message):
        ses = self.application.clone_session(self.session)
        return ses

    def set_session(self, message):
        self.check_experimenter()
        self.check_data(message)
        data = message['data']
        session = self.session
        if 'start_from_phase' in data and self.is_numeric(data['start_from_phase']):
            session.start_from_phase = int(data['start_from_phase'])
        if 'start_from_period' in data and self.is_numeric(data['start_from_period']):
            session.start_from_period = int(data['start_from_period'])
        if 'group_size' in data and self.is_numeric(data['group_size']):
            session.group_size = int(data['group_size'])
        if 'group_count' in data and self.is_numeric(data['group_count']):
            session.group_count = int(data['group_count'])
        if 'quantity_max' in data and self.is_numeric(data['quantity_max']):
            session.quantity_max = int(data['quantity_max'])
        if 'input_max' in data and self.is_numeric(data['input_max']):
            session.input_max = int(data['input_max'])
        if 'input_min' in data and self.is_numeric(data['input_min']):
            session.input_min = int(data['input_min'])
        if 'input_step_size' in data and self.is_numeric(data['input_step_size']):
            session.input_step_size = decimal.Decimal(str(data['input_step_size']))
        if 'input_step_time' in data and self.is_numeric(data['input_step_time']):
            session.input_step_time = int(data['input_step_time'])
        if 'cost_low' in data and self.is_numeric(data['cost_low']):
            session.cost_low = decimal.Decimal(str(data['cost_low']))
        if 'cost_high' in data and self.is_numeric(data['cost_high']):
            session.cost_high = decimal.Decimal(str(data['cost_high']))
        if 'starting_balance' in data and self.is_numeric(data['starting_balance']):
            session.starting_balance = decimal.Decimal(str(data['starting_balance']))
        if 'show_up_fee' in data and self.is_numeric(data['show_up_fee']):
            session.show_up_fee = decimal.Decimal(str(data['show_up_fee']))
        if 'exchange_rate' in data and self.is_numeric(data['exchange_rate']):
            session.exchange_rate = decimal.Decimal(str(data['exchange_rate']))
        if 'smallest_coin' in data and self.is_numeric(data['smallest_coin']):
            session.smallest_coin = decimal.Decimal(str(data['smallest_coin']))
        if 'currency' in data and data['currency'] in session.currencies:
            session.currency = data['currency']
        ses = self.application.clone_session(self.session)
        return ses

    def get_group(self, message):
        group = None
        key = None
        if self.session.is_started:
            if self.is_experimenter:
                self.check_data(message)
                key = message['data']
                period = self.application.get_current_period(self.session)
                group = period.groups[key]
            else:
                subject = self.application.get_subject(self.key)
                if 'data' in message and message['data'] != subject.group_key:
                    raise InvalidOperationException()
                group = subject.group
        return group

    def get_subject(self, message):
        subject = None
        key = None
        if self.is_experimenter:
            self.check_data(message)
            key = message['data']
        else:
            key = self.key
            subject = self.application.get_subject(key)
            if subject is None:
                subject = zook.Subject(self.session, self.key)
                self.application.set_subject(subject)
        return self.application.clone_subject(subject)

    def get_subjects(self, message):
        self.check_experimenter()
        ss = []
        for s in self.application.subjects.values():
            ss.append(s)
        return ss

    def set_subject(self, message):
        subject = None
        key = None
        if self.is_experimenter:
            self.check_data(message, ['key'])
            data = message['data']
            key = data['key']
        else:
            self.check_data(message)
            key = self.key
        data = message['data']
        subject = self.application.get_subject(key)
        if subject is None:
            subject = zook.Subject(self.session, key)
            self.application.set_subject(subject)
        if 'name' in data \
                and (self.is_experimenter or not subject.is_initialized):
            name = data['name']
            if not name or not name.strip():
                del self.application.subjects[key]
                raise InvalidOperationException('Client name can not be empty')
            s = self.application.get_subject(name)
            if s is not None:
                del self.application.subjects[key]
                raise InvalidOperationException('Client name already in use')
            subject.name = name
        elif not subject.name:
            raise InvalidOperationException('Client name can not be empty')
        if 'my_provide' in data:
            my_provide = data['my_provide'] or 0
            subject.my_provide = decimal.Decimal(my_provide)
        if 'my_bid' in data:
            my_bid = data['my_bid'] or 0
            subject.my_bid = decimal.Decimal(my_bid)
        if 'my_ask' in data:
            my_ask = data['my_ask'] or 0
            subject.my_ask = decimal.Decimal(my_ask)
        if 'real_name' in data:
            subject.real_name = data['real_name']
        if 'identification_number' in data:
            subject.identification_number = data['identification_number']
        if 'address' in data:
            subject.address = data['address']
        if 'postal_code' in data:
            subject.postal_code = data['postal_code']
        if 'location' in data:
            subject.location = data['location']
        if 'email' in data:
            subject.email = data['email']
        subject.decide_state()
        return subject

    def delete_subject(self, message):
        self.check_data(message)
        self.check_experimenter()
        key = message['data']
        socket = self.application.get_socket(key)
        subject = self.application.get_subject(key)
        if socket is not None and socket.is_open:
            socket.close()
        if subject is not None:
            del self.session.subjects[key]
            return True
        else:
            return False

    def suspend_subject(self, message):
        self.check_data(message)
        self.check_experimenter()
        key = message['data']
        subject = self.application.get_subject(key)
        if subject is not None:
            subject.is_suspended = True
            subject.set_state('robot')
            socket = self.application.get_socket(key)
            if socket is not None and socket.is_open:
                sub = self.application.clone_subject(subject)
                socket.send('get_subject', sub)
            self.application.proceed(self.session)
            return True
        else:
            return False

    def authorize(self, message):
        if self.is_experimenter:
            raise InvalidOperationException('You already are authorized')
        self.check_data(message, ['login', 'password'])
        data = message['data']
        login = data['login']
        password = data['password']
        if login == 'exp' and password == '1':
            self.is_experimenter = True
            e = zook.Experimenter(self.session, self.key)
            self.application.set_experimenter(e)
            return True
        else:
            raise UnauthorizedOperationException()

    def start_session(self, message):
        self.check_experimenter()
        ss = self.session.get_subjects_by_active()
        if len(ss) < 2:
            raise InvalidOperationException('There should be at least 2 active clients to start a session')
        self.application.start_session(self.session)
        ses = self.application.clone_session(self.session)
        return ses

    def pause_session(self, message):
        self.check_experimenter()
        if not self.session.is_started:
            raise InvalidOperationException('Session is not started')
        self.application.pause_session(self.session)
        ses = self.application.clone_session(self.session)
        return ses

    def resume_session(self, message):
        self.check_experimenter()
        if not self.session.is_started:
            raise InvalidOperationException('Session is not paused')
        self.application.resume_session(self.session)
        ses = self.application.clone_session(self.session)
        return ses

    def stop_session(self, message):
        self.check_experimenter()
        self.application.stop_session(self.session)
        ses = self.application.clone_session(self.session)
        return ses

    def continue_session(self, message):
        self.check_data(message)
        self.process_input(message)
        self.set_subject(message)
        if self.session.is_started:
            self.application.clear_timer(self.key)
            self.application.proceed(self.session)
        subject = self.application.get_subject(self.key)
        s = self.application.clone_subject(subject)
        return s

    def process_input(self, message):
        if not self.session.is_started:
            return
        subject = self.application.get_subject(self.key)
        group = subject.group
        data = message['data']
        if group.stage == 0:
            my_provide = None
            if 'my_provide' in data:
                my_provide = data['my_provide']
            if my_provide is None:
                raise InvalidOperationException(
                    'A default value will be used for you'
                    + ' unless you enter a valid number.')
            elif not self.is_numeric(my_provide) \
                    or float(my_provide) < 0 \
                    or float(my_provide) > 4 \
                    or (
                        float(my_provide) - float(my_provide) > 0
                        and float(my_provide) - float(my_provide) != 0.5
                        ):
                raise InvalidOperationException(
                    'The quantity must be at least 0,'
                    + ' at most 4, and a multiple of 1/2.'
                    )
        elif group.stage == 16:
            missing_fields = []
            real_name = None
            if 'real_name' in data:
                real_name = data['real_name']
            identification_number = None
            if 'identification_number' in data:
                identification_number = data['identification_number']
            address = None
            if 'address' in data:
                address = data['address']
            location = None
            if 'location' in data:
                location = data['location']
            email = None
            if 'email' in data:
                email = data['email']
            if not real_name or not real_name.strip():
                missing_fields.append('Full name')
            if not identification_number or not identification_number.strip():
                missing_fields.append('Identification Number')
            if not address or not address.strip():
                missing_fields.append('Address')
            if not location or not location.strip():
                missing_fields.append('Location')
            if not email or not email.strip():
                missing_fields.append('Email')
            if len(missing_fields) > 0:
                error = 'Please fill the mandatory fields listed below,\n\n'
                for f in missing_fields:
                    error += '- {0}\n'.format(f)
                raise InvalidOperationException(error)

    def skip_phase(self, message):
        self.check_experimenter()
        if self.session.is_started:
            self.session.phase.is_skipped = True
        ses = self.application.clone_session(self.session)
        return ses

    def reset(self, message):
        self.check_experimenter()
        ss = self.session.subjects.items()
        es = self.session.experimenters.items()
        self.application.sessions.clear()
        session = zook.Session()
        self.application.sessions[session.key] = session
        self.session = session
        e = self.application.experimenters[self.key]
        e.session = session
        session.experimenters[self.key] = e
        for key, s in ss:
            self.application.subjects.pop(key, None)
            socket = self.application.get_socket(key)
            if socket is not None and socket.is_open:
                socket.close()
        ses = self.application.clone_session(self.session)
        for key, e in es:
            if key is self.key:
                continue
            e.session = session
            session.experimenters[key] = e
            socket = self.application.get_socket(key)
            if socket is not None and socket.is_open:
                socket.send('get_session', ses)
        return ses